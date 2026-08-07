#!/usr/bin/env python3
"""
Contributor test script for SQL CDISC Rules Engine Contributor Repo.
"""

import sys
import argparse
import json
import logging
import warnings
import textwrap
import csv
import pandas as pd
import openpyxl as op
from glob import glob
from pathlib import Path
from typing import List, Optional, Tuple, Any, Dict

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
logging.basicConfig(level=logging.CRITICAL)

SDTM_RULES_DIR = Path("rules")
ADAM_RULES_DIR = Path("adam_rules")
ENGINE_DIR = Path("engine")


class ResultReporter:
    """Handles formatting, printing, and saving of test results."""

    @classmethod
    def json_to_readable(cls, results_data: dict, output_path: Path):
        """Write a human-readable summary of the JSON results to a text file."""
        with output_path.open("w") as f:
            if "error" in results_data:
                f.write("EXECUTION ERROR\n===============\n")
                f.write(f"{results_data['error']}\n")
                if "exception" in results_data:
                    f.write(f"Details: {results_data['exception']}\n")
                return

            datasets = results_data.get("datasets", [])
            if not datasets:
                f.write("No validation results found.\n")
                return

            total_errors = sum(len(ds.get("errors", [])) for ds in datasets)
            f.write(f"Total Errors Found: {total_errors}\n\n")

            for dataset in datasets:
                f.write(f"Dataset: {dataset.get('dataset', 'Unknown')}\n")
                f.write(f"Domain: {dataset.get('domain', 'N/A')}\n")
                if msg := dataset.get("execution_message"):
                    f.write(f"Rule Message: {msg}\n")

                errors = dataset.get("errors", [])
                f.write(f"Errors in this dataset: {len(errors)}\n")

                if not errors:
                    f.write("  No errors found in this dataset.\n")
                else:
                    f.write("\n")
                    for i, error in enumerate(errors, 1):
                        f.write(f"  Error {i}:\n")
                        if row := error.get("row"):
                            f.write(f"    Row: {row}\n")

                        for key, val in error.items():
                            if key not in ["row", "value", "validated"]:
                                f.write(f"    {key}: {val}\n")

                        if error.get("validated") is not None:
                            validation = "Yes" if error["validated"] else "No"
                            f.write(f"    Fully Validated in Test Case: {validation}\n")
                        f.write("\n")
                f.write("\n")

            if any(
                [
                    results_data.get("validated"),
                    results_data.get("unhighlighted_validations"),
                    results_data.get("unvalidated_highlights"),
                ]
            ):
                f.write("***ISSUES***\n")

            if results_data.get("validated") or results_data.get("unmatched_validation"):
                f.write("Mismatch between validation sheet and engine errors\n")
                f.write("Check results.json for more information, or examine the test case file directly.\n")

            if uh_v := results_data.get("unhighlighted_validations"):
                f.write("The following error groups on the validations sheet are not highlighted correctly:\n")
                f.write(", ".join(str(v) for e in uh_v for v in e.values()) + "\n")

            if uv_h := results_data.get("unvalidated_highlights"):
                f.write("The following rows have highlights that do not match the validations sheet:\n")
                ids = [f"{k}: {v}" for e in uv_h for k, v in e.items()]
                f.write("\n".join(", ".join(ids[i : i + 2]) for i in range(0, len(ids), 2)))

    @classmethod
    def save_case_results(
        cls, rules_dir: str, rule_id: str, test_type: str, case_id: str, results: dict, version_info: Optional[dict] = None
    ):
        """Saves JSON and TXT results to the file system."""
        base_rule_path = Path(rules_dir)
        rule_folder = base_rule_path / rule_id
        is_bundled = not rule_folder.exists()
        
        if is_bundled:
            for d in base_rule_path.iterdir():
                if d.is_dir() and (d / rule_id).exists():
                    rule_folder = d / rule_id
                    break

        if is_bundled:
            results_path = rule_folder / "results" / test_type / case_id
        else:
            results_path = rule_folder / test_type / case_id / "results"
            if not results_path.parent.exists():
                results_path = rule_folder / test_type / case_id 
                
        results_path.mkdir(parents=True, exist_ok=True)

        output = {**results, "dictionary_versions": version_info} if version_info else results
        with (results_path / "results.json").open("w") as f:
            json.dump(output, f, indent=2)

        cls.json_to_readable(results, results_path / "results.txt")
        return str(results_path)

    @staticmethod
    def display_rule_summary(summary: dict, verbose: bool = False):
        """Prints the execution summary for a single rule to the console."""
        print(f"\n{'='*60}\n{summary['rule_id']} Test Results Summary")
        print(f"\nRule: {summary['rule_id']}")
        print(f"Overall Status: {summary['status'].upper()}")

        for test_type in ["positive", "negative"]:
            tests = summary[f"{test_type}_tests"]
            print(f"{'-'*54}")
            print(f"{test_type.capitalize()} Test Cases: {len(tests)}")
            for test in tests:
                symbol = "[PASS]" if test["passed"] else "[FAIL]"
                print(f"\n  {symbol} Case {test['case_id']} - Results at: {test['results_path']}")

                if verbose:
                    txt_path = Path(test["results_path"]) / "results.txt"
                    if txt_path.exists():
                        print(f"\n{textwrap.indent(txt_path.read_text().strip(), "     ")}")

                if not test["passed"] and not verbose:
                    print(f"      Expected: {test['expected']}")
                    if test.get("total_errors") is not None:
                        print(f"      Got: {test['total_errors']} errors")
                    if test.get("error"):
                        print(f"      Error: {test['error']}")
                        print(f"      Exception: {test.get('exception', 'N/A')}")

        print("\n" + "=" * 60)


class TestRunner:
    """Test execution logic."""

    def __init__(
        self,
        standard: str,
        use_pgserver: bool = True,
        whodrug_path: Optional[str] = None,
        meddra_path: Optional[str] = None,
        unii_path: Optional[str] = None,
        medrt_path: Optional[str] = None,
        loinc_path: Optional[str] = None,
        snomed_path: Optional[str] = None,
        ct: Optional[str] = None,
    ):
        from dotenv import load_dotenv

        load_dotenv("engine/.env.example")
        self._setup_engine_path()
        self.use_pgserver = use_pgserver
        self.standard = standard
        self.rules_dir = None
        if standard == "sdtm":
            self.rules_dir = SDTM_RULES_DIR
        elif standard == "adam":
            self.rules_dir = ADAM_RULES_DIR
        else:
            raise ValueError(f"Unsupported standard: {standard}")

        from engine.cdisc_rules_engine.models.sql_external_dictionaries_container import (
            SqlExternalDictionariesContainer,
        )

        ext_dicts = SqlExternalDictionariesContainer(
            dictionary_path_mapping={
                "whodrug": whodrug_path,
                "meddra": meddra_path,
                "unii": unii_path if unii_path != "default" else "dummy_ex_dicts/unii",
                "medrt": medrt_path if medrt_path != "default" else "dummy_ex_dicts/medrt",
                "loinc": loinc_path if loinc_path != "default" else "dummy_ex_dicts/loinc",
                "snomed": snomed_path if snomed_path != "default" else "dummy_ex_dicts/snomed",
            }
        )

        self.version_info = self.get_ext_dict_versions(ext_dicts) if ext_dicts else {}

        from engine.cdisc_rules_engine.data_service.postgresql_data_service import PostgresQLDataService

        cache_dir = "resources/cache"
        latest_codelist_path = max(glob(f"engine/{cache_dir}/{self.standard}ct-[0-9][0-9][0-9][0-9]-*.pkl"), default=None)
        latest_codelist_file = Path(latest_codelist_path).name if latest_codelist_path else None

        self.data_service = PostgresQLDataService.instance(
            use_pgserver=self.use_pgserver,
            codelists=[latest_codelist_file] if latest_codelist_file else [],
            provided_codelists=ct,
            cache_path=cache_dir,
            external_dictionaries=ext_dicts,
        )

    @staticmethod
    def _setup_engine_path():
        """Ensures the engine submodule is in sys.path."""
        if str(ENGINE_DIR) not in sys.path:
            sys.path.insert(0, str(ENGINE_DIR))

    @staticmethod
    def get_ext_dict_versions(ext_dicts):
        import importlib
        import dataclasses

        READER_MAP = {
            "meddra": ("engine.cdisc_rules_engine.readers.external_dictionary_readers.meddra_reader", "MeddraReader"),
            "whodrug": (
                "engine.cdisc_rules_engine.readers.external_dictionary_readers.whodrug_reader",
                "WhoDrugReader",
            ),
            "loinc": ("engine.cdisc_rules_engine.readers.external_dictionary_readers.loinc_reader", "LoincReader"),
            "unii": ("engine.cdisc_rules_engine.readers.external_dictionary_readers.unii_reader", "UniiReader"),
            "medrt": ("engine.cdisc_rules_engine.readers.external_dictionary_readers.medrt_reader", "MedRTReader"),
            "snomed": ("engine.cdisc_rules_engine.readers.external_dictionary_readers.snomed_reader", "SnomedReader"),
        }

        version_info = {}

        for dict_type, path in ext_dicts.dictionary_path_mapping.items():
            if path and dict_type in READER_MAP:
                module_path, class_name = READER_MAP[dict_type]
                reader_cls = getattr(importlib.import_module(module_path), class_name)
                reader = reader_cls(pgi=None, dictionary_path=path)
                version_info[dict_type] = dataclasses.asdict(reader._extract_version_metadata())

        return version_info

    def get_rule_path(self, rule_id: str) -> Optional[Path]:
        if not self.rules_dir.exists():
            return None
            
        standalone_path = self.rules_dir / rule_id
        if standalone_path.exists() and standalone_path.is_dir():
            return standalone_path
            
        for d in self.rules_dir.iterdir():
            if d.is_dir():
                bundled_path = d / rule_id
                if bundled_path.exists() and bundled_path.is_dir():
                    return bundled_path
                    
        return None

    def get_available_rules(self) -> List[str]:
        if not self.rules_dir.exists():
            return []
            
        rules = []
        for d in self.rules_dir.iterdir():
            if not d.is_dir(): 
                continue
            
            if (d.name.startswith("CORE-") or d.name.startswith("AD") or d.name.startswith("NEW-RULE")):
                if list(d.glob("[!~]*.yml")):
                    rules.append(d.name)
                else:
                    for subd in d.iterdir():
                        if subd.is_dir() and list(subd.glob("[!~]*.yml")):
                            rules.append(subd.name)
                            
        return sorted(rules)

    def get_test_cases(self, rule_id: str) -> dict:
        """Scans directories to find available test cases for a rule."""
        cases = {"positive": [], "negative": []}
        rule_path = self.get_rule_path(rule_id)
        
        if not rule_path:
            return cases

        parent_dir = rule_path.parent
        shared_cases_path = parent_dir / "shared_test_cases"
        
        test_source = shared_cases_path if shared_cases_path.exists() else rule_path

        for test_type in ["positive", "negative"]:
            test_type_path = test_source / test_type
            if test_type_path.exists():
                for case_dir in sorted(test_type_path.iterdir()):
                    if case_dir.is_dir():
                        data_dir = case_dir / "data"
                        if not data_dir.exists():
                            data_dir = case_dir
                            
                        has_csv = list(data_dir.glob("*.csv"))
                        has_xlsx = list(data_dir.glob("[!~]*.xls*"))
                        
                        if has_csv:
                            cases[test_type].append({"case_id": case_dir.name, "data_path": str(data_dir), "format": "csv"})
                        elif has_xlsx:
                            cases[test_type].append({"case_id": case_dir.name, "data_path": str(data_dir), "format": "xlsx"})
                            
        return cases

    @staticmethod
    def _read_library_specs_csv(data_path: str) -> Tuple[str, str, List[str]]:
        lib_path = Path(data_path) / "_library.csv"
        if not lib_path.exists():
            raise ValueError(f"_library.csv not found in {data_path}")

        with lib_path.open("r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers = next(reader)

            try:
                first_row = next(reader)
            except StopIteration:
                raise ValueError(f"library.csv in {data_path} is empty")

            standard = str(first_row[0]).strip()
            version = str(first_row[1]).strip().replace("-", ".")
            ct_list = [f"{str(row[0]).strip()}-{str(row[1]).strip()}" for row in reader if len(row) >= 2 and row[0] and str(row[0]).strip().lower().endswith("ct")]
            return standard, version, ct_list

    @staticmethod
    def _read_library_specs_xlsx(excel_path: str) -> Tuple[str, str, List[str]]:
        wb = op.load_workbook(excel_path, data_only=True, read_only=True)
        try:
            if "Library" not in wb.sheetnames:
                raise ValueError(f"Sheet 'Library' not found in {excel_path}")
            ws = wb["Library"]
            row_iter = ws.iter_rows(min_row=2, max_col=2, values_only=True)
            first_row = next(row_iter)
            standard = str(first_row[0]).strip()
            version = str(first_row[1]).strip().replace("-", ".")
            ct_list = [f"{str(row_data[0]).strip()}-{str(row_data[1]).strip()}" for row_data in row_iter if row_data[0] and str(row_data[0]).strip().lower().endswith("ct") and row_data[1]]
            return standard, version, ct_list
        finally:
            wb.close()

    def _load_csv_datasets(self, data_path: str) -> list:
        from engine.tests.rule_regression.regression import TestDataset, VariableMetadata
        datasets_csv_path = Path(data_path) / "_datasets.csv"
        datasets_df = pd.read_csv(datasets_csv_path)
        test_datasets = []

        for _, row in datasets_df.iterrows():
            filename = row["Filename"]
            label = row.get("Label", "")
            dataset_path = Path(data_path) / filename

            if dataset_path.exists():
                dataset_df = pd.read_csv(dataset_path, keep_default_na=False, na_values=[""])
                variables = []
                col_type_dict = {}

                for i, col in enumerate(dataset_df.columns):
                    if col.startswith("Unnamed:"): continue
                    var_type = str(dataset_df[col].iloc[1])
                    variables.append(VariableMetadata(name=col, label=str(dataset_df[col].iloc[0]), type=var_type, length=dataset_df[col].iloc[2], format="", order=i + 1))
                    col_type_dict[col] = var_type

                data = {}
                for col in dataset_df.columns:
                    if col.startswith("Unnamed:"): continue
                    column_values = dataset_df[col].iloc[3:].tolist()
                    if col_type_dict[col].lower() == "num":
                        column_values = [None if pd.isna(val) or val == "" else float(val) for val in column_values]
                    elif col_type_dict[col].lower() == "char":
                        column_values = ["" if pd.isna(val) else str(val) for val in column_values]
                    data[col] = column_values

                test_datasets.append(TestDataset(filename=filename, name=filename.split(".")[0].upper(), label=label, variables=variables, records=data))
        return test_datasets

    @staticmethod
    def _init_engine_specs(standard: str, standard_version: str):
        try:
            from engine.cdisc_rules_engine.utilities.ig_specification import IGSpecification
            return IGSpecification(standard=standard, standard_version=standard_version, standard_substandard=None, define_xml_version=None)
        except ImportError:
            print("Error: Could not import engine modules. Is the submodule initialised?")
            sys.exit(1)

    def run_validation(self, rule_id: str, case_info: dict) -> Tuple[Any, Optional[dict]]:
        rule_path = self.get_rule_path(rule_id)
        if not rule_path:
            return None, {"error": "Rule path missing", "exception": f"Could not find path for rule {rule_id}"}
        
        rule_ymls = list(rule_path.glob("[!~]*.yml"))
        if not rule_ymls:
            return None, {"error": "Rule YAML missing", "exception": f"No YAML found in {rule_path}"}
        
        data_path_obj = Path(case_info["data_path"])
        is_csv = case_info["format"] == "csv"

        for file in ["define.xml", "stf.xml"]:
            file_path = str(data_path_obj / file) if (data_path_obj / file).exists() else str(rule_path / file) if (rule_path / file).exists() else None
            if file == "define.xml":
                define_xml_path = file_path
                self.data_service._update_define_xml_path(define_xml_path)
                if define_xml_path:
                    from engine.cdisc_rules_engine.services.define_xml.define_xml_reader_factory import DefineXMLReaderFactory
                    extensible_terms = DefineXMLReaderFactory.from_filename(define_xml_path).get_extensible_codelist_mappings()
                    self.data_service._add_extensible_ct_terms(extensible_terms)
            elif file == "stf.xml":
                self.data_service._update_stf_file_path(file_path)

        try:
            import yaml
            from engine.tests.rule_regression.regression import process_test_case_dataset_sql, sharepoint_xlsx_to_test_datasets

            with open(rule_ymls[0], "r", encoding="utf-8") as f:
                rule = yaml.safe_load(f)

            if is_csv:
                standard, standard_version, provided_codelists = self._read_library_specs_csv(str(data_path_obj))
                test_datasets = self._load_csv_datasets(str(data_path_obj))
            else:
                excel_file = (list(data_path_obj.glob("[!~]*.xlsx")) + list(data_path_obj.glob("[!~]*.xls")))[0]
                standard, standard_version, provided_codelists = self._read_library_specs_xlsx(str(excel_file))
                test_datasets = sharepoint_xlsx_to_test_datasets(str(excel_file))

            ig_specs = self._init_engine_specs(standard, standard_version)
            if provided_codelists:
                self.data_service._update_provided_codelists(provided_codelists)

            sql_results, sql_regression = process_test_case_dataset_sql(
                regression_errors={}, define_xml_file_path=define_xml_path, data_test_datasets=test_datasets,
                ig_specs=ig_specs, rule=rule, use_pgserver=self.use_pgserver, data_service=self.data_service
            )

            return sql_results, {"datasets": sql_regression} if sql_regression else {"datasets": []}
        except Exception as e:
            return None, {"error": "Error executing engine validation.", "exception": str(e)}

    def evaluate_case(self, rule_id: str, test_type: str, case_info: dict) -> dict:
        case_id = case_info["case_id"]
        is_csv = case_info["format"] == "csv"
        expected = "0 errors" if test_type == "positive" else ">0 errors"

        _, results_data = self.run_validation(rule_id, case_info)

        if results_data is None or results_data.get("error"):
            results_path = ResultReporter.save_case_results(self.rules_dir, rule_id, test_type, case_id, results_data or {"error": "Unknown", "exception": "Engine returned None"}, self.version_info)
            return {"case_id": case_id, "passed": False, "total_errors": None, "expected": expected, "error": results_data.get("error"), "results_path": results_path}

        if test_type == "negative":
            validations = self.get_validation_info_csv(case_info["data_path"]) if is_csv else self.get_validation_info_xlsx(case_info["data_path"])
            results_data, unmatched = self.validate_errors(results_data, validations, rule_id if is_csv else None)
            
            if unmatched:
                results_data["unmatched_validation"] = unmatched

            if not is_csv:
                highlights = self.get_excel_highlights(case_info["data_path"])
                unhighlighted_validations, unvalidated_highlights = self.check_highlights(validations, highlights)
                if unhighlighted_validations: results_data["unhighlighted_validations"] = unhighlighted_validations
                if unvalidated_highlights: results_data["unvalidated_highlights"] = unvalidated_highlights

        results_path = ResultReporter.save_case_results(self.rules_dir, rule_id, test_type, case_id, results_data, self.version_info)
        total_errors = sum(len(ds.get("errors", [])) for ds in results_data.get("datasets", []))
        passed = (total_errors == 0) if test_type == "positive" else (total_errors > 0)

        return {"case_id": case_id, "passed": passed, "total_errors": total_errors, "expected": expected, "results_path": results_path}

    def get_validation_info_csv(self, data_path: str) -> dict:
        validation_values = {}
        val_path = Path(data_path) / "_validation.csv"
        if not val_path.exists(): return validation_values

        with val_path.open("r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            try: headers = next(reader)
            except StopIteration: return validation_values

            rule_id_idx = headers.index("Rule ID") if "Rule ID" in headers else -1
            error_group_idx = headers.index("Error group") if "Error group" in headers else -1

            for row in reader:
                if not row or not any(row): continue
                r_id = row[rule_id_idx].strip() if rule_id_idx != -1 and len(row) > rule_id_idx else None
                e_id = row[error_group_idx].strip() if error_group_idx != -1 and len(row) > error_group_idx else "1"
                if len(row) == len(headers):
                    validation_values.setdefault(r_id, {}).setdefault(e_id, []).append(dict(zip(headers, row)))
        return validation_values

    def get_validation_info_xlsx(self, data_path: str):
        xl_path = list(Path(data_path).glob("[!~]*.xls*"))[0]
        wb = op.load_workbook(xl_path, data_only=True)
        validation_values = {}
        if "Validation" in wb.sheetnames:
            ws = wb["Validation"]
            headers = [cell.value for cell in ws[1]][1:]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None: continue
                validation_values.setdefault(None, {}).setdefault(row[0], []).append(dict(zip(headers, row[1:])))
        return validation_values

    def get_excel_highlights(self, data_path: str):
        xl_path = list(Path(data_path).glob("[!~]*.xls*"))[0]
        highlighted_cells = {}
        YELLOW_INDICES = (5, 11, 13, 14, 34)
        wb = op.load_workbook(xl_path, data_only=True)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    fg = cell.fill.fgColor
                    if not fg: continue
                    is_yellow = (isinstance(fg.rgb, str) and fg.rgb.lower().endswith("ffff00")) or (isinstance(fg.indexed, int) and fg.indexed in YELLOW_INDICES)
                    if is_yellow:
                        highlighted_cells.setdefault(sheet.title, {}).setdefault(int(cell.row), {}).update({sheet.cell(row=1, column=cell.column).value: cell.value})
        return highlighted_cells

    def validate_errors(self, results_data: dict, validations: dict, rule_id: str = None):
        unmatched = []
        rule_validations = validations.get(rule_id.split("/")[-1] if rule_id else None, validations.get(None, {}))
        
        flat_validation = {}
        for idx, entries in rule_validations.items():
            if not entries: continue
            error_level = entries[0].get("Error level", "Record").lower()
            sheet = entries[0]["Sheet"]
            row = 1 if str(entries[0]["Row num"]).lower() in ["1", "n/a"] else int(entries[0]["Row num"]) - 4
            flat_validation[(sheet, error_level, row, idx)] = {e["Variable"]: e["Error value"] for e in entries}

        for ds in results_data.get("datasets", []):
            sheet_name = ds["dataset"].split(".")[0].lower() if rule_id else ds["dataset"]
            for error_obj in ds.get("errors", []):
                if not error_obj.get("error"):
                    match_found = False
                    for (v_sheet, v_error_level, v_row, v_idx), v_values in flat_validation.items():
                        if v_sheet.split(".")[0].lower() == ds["dataset"].split(".")[0].lower():
                            v_not_absent = set(k for k, v in v_values.items() if v != "[ABSENT]")
                            res_not_absent = set(k for k, v in error_obj["value"].items() if v != "[ABSENT]") - set(k for k, v in v_values.items() if v == "[ABSENT]")
                            if v_not_absent == res_not_absent or not res_not_absent:
                                match_found = True
                            if match_found:
                                error_obj["validated"] = True
                                del flat_validation[(v_sheet, v_error_level, v_row, v_idx)]
                                break
                    if not match_found: error_obj["validated"] = False

        for _, values in flat_validation.items():
            unmatched.append({"value": dict(values)})
        return results_data, unmatched

    def check_highlights(self, validations: dict, highlights: dict):
        unmatched_validations, unmatched_highlights, matched_highlights = [], [], set()
        
        for v_entries in validations.get(None, {}).values():
            for e in v_entries:
                sheet, error_level, row, var, error_val = e["Sheet"], e["Error level"].lower(), e["Row num"], e["Variable"], e["Error value"]
                if str(var)[0] == "$": continue
                var = var.split(".")[-1] if "." in str(var) else var
                if error_level == "record" and error_val != "[ABSENT]":
                    if str(highlights.get(sheet, {}).get(row, {}).get(var)) == str(error_val): matched_highlights.add((sheet, row, var))
                    else: unmatched_validations.append({None: [sheet, error_level, row, var, error_val]})
                elif error_level == "variable" and error_val == "[PRESENT]":
                    if var in highlights.get(sheet, {}).get(row, {}): matched_highlights.add((sheet, row, var))
                    else: unmatched_validations.append({None: [sheet, error_level, row, var, error_val]})

        for sheet, rows in highlights.items():
            for row_num, vals in rows.items():
                for var in vals.keys():
                    if (sheet, row_num, var) not in matched_highlights:
                        unmatched_highlights.append({"Sheet": sheet, "Row": row_num, "Variable": var})

        return unmatched_validations, unmatched_highlights

    def _get_cases_to_run(self, rule_id: str, specific_case: str = None) -> Dict[str, List[dict]]:
        all_cases = self.get_test_cases(rule_id)
        if not specific_case: return all_cases

        target_type, target_id = specific_case.split("/")
        filtered = {"positive": [], "negative": []}
        if target_type in filtered:
            if found := next((c for c in all_cases[target_type] if c["case_id"] == target_id), None):
                filtered[target_type].append(found)
        return filtered

    def run_rule_suite(self, rule_id: str, specific_case: str = None) -> dict:
        summary = {"rule_id": rule_id, "positive_tests": [], "negative_tests": [], "status": "passed"}
        cases_to_run = self._get_cases_to_run(rule_id, specific_case)

        if not cases_to_run["positive"] and not cases_to_run["negative"]: return summary

        for test_type in ["positive", "negative"]:
            for case in cases_to_run[test_type]:
                result = self.evaluate_case(rule_id, test_type, case)
                summary[f"{test_type}_tests"].append(result)
                if not result["passed"]: summary["status"] = "failed"

        return summary


class InteractiveHandler:
    """Handles user prompts for interactive mode."""

    @staticmethod
    def prompt_rule(available: List[str]) -> str:
        print("\nWhich rule would you like to test?")
        while True:
            choice = input("Enter rule ID (e.g. CORE-000215): ").strip()
            if choice in available: return choice
            print(f"Invalid. Available: {', '.join(available[:5])}...")

    @staticmethod
    def prompt_case(available: dict) -> Optional[str]:
        print("\nTest specific case? (Leave blank for all)")
        flat_list = [f"{t_type}/{c['case_id']}" for t_type in ["positive", "negative"] for c in available[t_type]]
        if not flat_list: return None

        for i, tc in enumerate(flat_list, 1): print(f"  {i}. {tc}")
        while True:
            choice = input("\nEnter case (e.g., positive/01, number, or Enter): ").strip()
            if not choice: return None
            if choice in flat_list: return choice
            if choice.isdigit() and 0 <= int(choice) - 1 < len(flat_list): return flat_list[int(choice) - 1]
            print("Invalid choice.")


def parse_args():
    parser = argparse.ArgumentParser(description="CDISC SQL Rules Engine Tester")
    parser.add_argument("-s", "--standard", help="Provide the standard (e.g. sdtm, adam) to test against.", default="sdtm")
    parser.add_argument("-r", "--rule", help="Rule ID (e.g., CORE-000176 or NEW-RULE)")
    parser.add_argument("-all", "--all-rules", action="store_true", help="Run all rules")
    parser.add_argument("-tc", "--test-case", help="Specific case (e.g., positive/01)")
    parser.add_argument("-v", "--verbose", action="store_true", help="Print detailed results")
    parser.add_argument("-pg", "--use-postgres", action="store_true", help="Use standard PostgreSQL instead of pgserver default")
    parser.add_argument("-wd", "--whodrug", help="Provide path to WHODrug files")
    parser.add_argument("-md", "--meddra", help="Provide path to MedDRA files")
    parser.add_argument("-un", "--unii", help="Provide path to UNII files")
    parser.add_argument("-mrt", "--medrt", help="Provide path to Med-RT files")
    parser.add_argument("-lo", "--loinc", help="Provide path to LOINC files")
    parser.add_argument("-sno", "--snomed", help="Provide path to SNOMED files")
    return parser.parse_args()


def main():
    args = parse_args()
    runner = TestRunner(
        standard=args.standard.lower(),
        use_pgserver=not args.use_postgres,
        whodrug_path=args.whodrug, meddra_path=args.meddra, unii_path=args.unii,
        medrt_path=args.medrt, loinc_path=args.loinc, snomed_path=args.snomed,
    )
    available_rules = runner.get_available_rules()

    if not available_rules:
        print("Error: No rules found in 'rules' directory.")
        sys.exit(1)

    rules_to_run = available_rules
    specific_case = None

    if not args.all_rules:
        rule_input = args.rule or InteractiveHandler.prompt_rule(available_rules)
        
        bundle_path = runner.rules_dir / rule_input
        if bundle_path.is_dir() and (bundle_path / "shared_test_cases").exists():
            rules_to_run = sorted([d.name for d in bundle_path.iterdir() if d.is_dir() and d.name != "shared_test_cases"])
            print(f"\nIdentified {rule_input} as a bundle. Running all {len(rules_to_run)} rules inside it.")
        elif rule_input in available_rules:
            rules_to_run = [rule_input]
            specific_case = args.test_case if args.rule and args.test_case else InteractiveHandler.prompt_case(runner.get_test_cases(rule_input)) if not args.rule else None
        else:
            print(f"Error: {rule_input} is not a valid rule or bundle.")
            sys.exit(1)
            
        if len(rules_to_run) == 1:
            print(f"\nRunning {rules_to_run[0]}...")
            summary = runner.run_rule_suite(rules_to_run[0], specific_case)
            ResultReporter.display_rule_summary(summary, verbose=args.verbose)
            sys.exit(0 if summary["status"] == "passed" else 1)

    if args.test_case and args.all_rules:
        print("Error: --test-case cannot be used with --all-rules.")
        sys.exit(1)

    results = {"passed": [], "failed": [], "error": []}
    total = len(rules_to_run)
    print("Core SQL Rules Engine - Test Suite\n" + "=" * 60)

    for i, rule_id in enumerate(rules_to_run, 1):
        sys.stdout.write(f"\r[{i}/{total}] Testing {rule_id}...")
        sys.stdout.flush()
        summary = runner.run_rule_suite(rule_id)
        if any(t.get("error") for t in summary["positive_tests"] + summary["negative_tests"]): results["error"].append(summary)
        elif summary["status"] == "passed": results["passed"].append(summary)
        else: results["failed"].append(summary)

    sys.stdout.write("\n\n")
    print("=" * 60 + "\nFINAL SUMMARY\n" + "=" * 60)
    print(f"Total: {total} | Passed: {len(results['passed'])} | Failed: {len(results['failed'])} | Errors: {len(results['error'])}")

    if results["failed"]:
        print("\nFailed Validation:")
        for s in results["failed"]:
            print(f"  - {s['rule_id']}")
            if args.verbose:
                for t in s["positive_tests"] + s["negative_tests"]:
                    if not t["passed"]: print(f"      Case {t['case_id']}: Expected {t['expected']}, Got {t['total_errors']} errors")

    if results["error"]:
        print("\nExecution Errors:")
        for s in results["error"]:
            print(f"  - {s['rule_id']}")
            if args.verbose:
                for t in s["positive_tests"] + s["negative_tests"]:
                    if t.get("error"):
                        print(f"      Case {t['case_id']}\n      - Error: {t['error']}" + (f"\n      - Exception: {t['exception']}" if t.get("exception") else ""))

    sys.exit(1 if results["failed"] or results["error"] else 0)


def generate_rule_results(standard: str, rules_dir: str, rule_id: str) -> dict:
    rule_yml = list((Path(rules_dir) / rule_id).glob("[!~]*.yml"))[0]
    with rule_yml.open("r", encoding="utf-8") as f:
        content = f.read().lower()
    
    return TestRunner(
        standard=standard,
        unii_path="dummy_ex_dicts/unii" if "unii" in content else None,
        medrt_path="dummy_ex_dicts/medrt" if "medrt" in content else None,
        loinc_path="dummy_ex_dicts/loinc" if "loinc" in content else None,
        snomed_path="dummy_ex_dicts/snomed" if "snomed" in content else None,
    ).run_rule_suite(rule_id)


if __name__ == "__main__":
    try: main()
    except KeyboardInterrupt: sys.exit(1)