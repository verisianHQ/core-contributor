"""
Script to create a new rule directory with the required structure and template files for testing.
"""

import sys
import shutil
import re
import json
import csv
from pathlib import Path
import openpyxl

SDTM_RULES_DIR = Path("rules")
ADAM_RULES_DIR = Path("adam_rules")
PLACEHOLDER_RULE_ID = "NEW-RULE"

def create_excel_file(filepath: Path, is_negative: bool, is_bundled: bool):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_lib = wb.create_sheet("Library")
    ws_lib.append(["Product", "Version"])
    ws_lib["A2"] = "sdtmig"
    ws_lib["B2"] = "3-4"

    ws_data = wb.create_sheet("Datasets")
    ws_data.append(["Filename", "Label"])
    ws_data.append(["", ""])

    if is_negative:
        ws_val = wb.create_sheet("Validation")
        headers = ["Error group", "Sheet", "Error level", "Row num", "Variable", "Error value"]
        
        if is_bundled:
            headers.insert(0, "Rule ID")
            
        ws_val.append(headers)
        ws_val.append([""] * len(headers))

    wb.save(filepath)

def create_csv_files(case_dir: Path, is_negative: bool, is_bundled: bool):
    with (case_dir / "library.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Product", "Version"])
        writer.writerow(["sdtmig", "3-4"])

    with (case_dir / "datasets.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Filename", "Label"])
        writer.writerow(["", ""])

    if is_negative:
        with (case_dir / "validation.csv").open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            headers = ["Error group", "Sheet", "Error level", "Row num", "Variable", "Error value"]
            
            if is_bundled:
                headers.insert(0, "Rule ID")
                
            writer.writerow(headers)
            writer.writerow([""] * len(headers))

def create_test_cases(base_dir: Path, test_type: str, count: int, is_bundled: bool, format_choice: str, file_prefix: str):
    for i in range(1, count + 1):
        case_id = f"{i:02d}"
        
        case_dir = base_dir / test_type / case_id
        case_dir.mkdir(parents=True, exist_ok=True)
        
        if format_choice == 'csv':
            create_csv_files(case_dir, is_negative=(test_type == "negative"), is_bundled=is_bundled)
        else:
            excel_filename = f"{file_prefix.lower()}-{test_type}-{case_id}.xlsx"
            excel_path = case_dir / excel_filename
            create_excel_file(excel_path, is_negative=(test_type == "negative"), is_bundled=is_bundled)

def main():
    # user input for which rules dir
    rules_choice = input("Which rules directory would you like to use? (sdtm/adam) ").lower()
    if rules_choice == "sdtm":
        RULES_DIR = SDTM_RULES_DIR
    elif rules_choice == "adam":
        RULES_DIR = ADAM_RULES_DIR
    else:
        print("Invalid choice. Aborting.")
        sys.exit(1)

    is_bundled = input("Is this part of a bundled rule structure sharing test cases? (y/n) ").lower() == 'y'
    
    if is_bundled:
        bundle_name = input("Enter the bundle folder name (e.g., AD0020-32): ").strip()
        bundle_dir = RULES_DIR / bundle_name
        test_cases_dir = bundle_dir / "shared_test_cases"
        
        range_input = input("Enter the rule ID range (e.g., AD0020-AD0032): ").strip()
        
        match = re.match(r"([A-Za-z]+)(\d+)-([A-Za-z]*)(\d+)", range_input)
        if not match:
            print("Invalid range format. Please use format like AD0020-AD0032. Aborting.")
            sys.exit(1)
            
        prefix1, num1, _, num2 = match.groups()
        pad = len(num1)
        start_num, end_num = int(num1), int(num2)
        prefix = prefix1.upper()
        
        rule_ids = [f"{prefix}{i:0{pad}d}" for i in range(start_num, end_num + 1)]
        
        bundle_dir.mkdir(parents=True, exist_ok=True)
        test_cases_dir.mkdir(parents=True, exist_ok=True)
        
        for r_id in rule_ids:
            rule_dir = bundle_dir / r_id
            if rule_dir.exists():
                shutil.rmtree(rule_dir, ignore_errors=True)
            rule_dir.mkdir(parents=True, exist_ok=True)
            
            yml_file = rule_dir / f"{r_id}.yml"
            try:
                shutil.copy("tests/template-rule.yml", yml_file)
            except FileNotFoundError:
                yml_file.touch()
                
            res_dir = rule_dir / "results"
            res_dir.mkdir(parents=True, exist_ok=True)
            with (res_dir / "results.json").open("w") as f:
                json.dump({}, f)
            (res_dir / "results.txt").touch()

        file_prefix = bundle_name

    else:
        rule_dir = RULES_DIR / PLACEHOLDER_RULE_ID
        test_cases_dir = rule_dir

        if rule_dir.exists():
            do_wipe = input("Another NEW-RULE folder already exists here. Erase it and make a new one? (y/n) ").lower() == "y"
            if not do_wipe:
                print("Aborting.")
                sys.exit(0)
            shutil.rmtree(rule_dir, ignore_errors=False)

        rule_dir.mkdir(parents=True, exist_ok=True)

        yml_file = rule_dir / f"{PLACEHOLDER_RULE_ID.lower()}.yml"
        try:
            shutil.copy("tests/template-rule.yml", yml_file)
        except FileNotFoundError:
            yml_file.touch()
            
        file_prefix = PLACEHOLDER_RULE_ID

    format_choice = ""
    while format_choice not in ["csv", "xlsx"]:
        format_choice = input("Would you like test cases in CSV or XLSX format? (csv/xlsx): ").strip().lower()
        if format_choice not in ["csv", "xlsx"]:
            raise ValueError("Invalid format. Please enter 'csv' or 'xlsx'.")

    n_pos_cases = int(input("Enter the number of positive test cases to create: "))
    n_neg_cases = int(input("Enter the number of negative test cases to create: "))

    if n_pos_cases > 0:
        create_test_cases(test_cases_dir, "positive", n_pos_cases, is_bundled, format_choice, file_prefix)

    if n_neg_cases > 0:
        create_test_cases(test_cases_dir, "negative", n_neg_cases, is_bundled, format_choice, file_prefix)

    print(f"\nSuccess!")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted.")
        sys.exit(1)