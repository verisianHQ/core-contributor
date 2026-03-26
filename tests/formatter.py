from copy import copy
import sys
import argparse
from pathlib import Path
from typing import List, Optional, Dict
import openpyxl
from openpyxl.styles import Font, PatternFill

RULES_DIR = Path("rules")

BOLD_FONT = Font(bold=True)
ITALIC_FONT = Font(italic=True)
LIGHT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

LIBRARY_SHEET = {
    "name": "Library",
    "index": 0,
    "headers": ["Product", "Version"],
    "defaults": {"A2": "sdtmig", "B2": "3-4"},
}

DATASETS_SHEET = {
    "name": "Datasets",
    "index": 1,
    "headers": ["Filename", "Label"],
}

VALIDATION_SHEET = {
    "name": "Validation",
    "index": 2,
    "headers": ["Error group", "Sheet", "Error level", "Row num", "Variable", "Error value"],
}


class FileManager:
    """Handles discovery of test case files."""

    @staticmethod
    def get_all_test_cases() -> List[dict]:
        """Returns a flat list of all test cases across all rules."""
        if not RULES_DIR.exists():
            return []

        cases = []
        rule_dirs = sorted([d for d in RULES_DIR.iterdir() if d.is_dir() and d.name.startswith("CORE-")])

        for rule_path in rule_dirs:
            for test_type in ["positive", "negative"]:
                type_dir = rule_path / test_type
                if not type_dir.exists():
                    continue

                for case_dir in sorted(type_dir.iterdir()):
                    if not case_dir.is_dir():
                        continue

                    data_dir = case_dir / "data"
                    excel_files = list(data_dir.glob("[!~]*.xls*"))  # Ignore temp files

                    if excel_files:
                        cases.append(
                            {
                                "rule_id": rule_path.name,
                                "case_id": case_dir.name,
                                "type": test_type,
                                "path": excel_files[0],
                            }
                        )
        return cases


class LabelManager:
    """Scrapes existing sheets to build a global map of Filename -> Label."""

    def __init__(self):
        self.label_map: Dict[str, str] = {}

    def scan_all_files(self, cases: List[dict]):
        print("Scanning files to build label map...")
        for case in cases:
            self._scan_file(case["path"])
        print(f"Map built. Found {len(self.label_map)} unique labels.")

    def _scan_file(self, path: Path):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            if "Datasets" in wb.sheetnames:
                ws = wb["Datasets"]
                for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                    if row and len(row) >= 2 and row[0] and row[1]:
                        self.label_map[str(row[0])] = str(row[1])
            wb.close()
        except Exception:
            pass

    def get_label(self, filename: str) -> str:
        return self.label_map.get(filename, "Unknown")


class Formatter:
    """
    Orchestrates the formatting of a single Excel file.
    """

    def __init__(self, file_path: Path, label_manager: LabelManager):
        self.file_path = file_path
        self.label_manager = label_manager
        self.wb: Optional[openpyxl.Workbook] = None
        self.is_negative = self.file_path.parents[2].name == "negative"

    def format(self) -> bool:
        """Main execution flow."""
        try:
            self.wb = openpyxl.load_workbook(self.file_path)

            self._ensure_sheets()
            self._populate_sheets()
            self._style_sheets()

            self.wb.save(self.file_path)
            self.wb.close()
            return True
        except Exception as e:
            print(f"Failed to format {self.file_path}: {e}")
            return False

    def _ensure_sheets(self):
        """Creates missing sheets and orders tabs."""
        self._ensure_sheet_exists(LIBRARY_SHEET["name"], LIBRARY_SHEET["index"])
        self._ensure_sheet_exists(DATASETS_SHEET["name"], DATASETS_SHEET["index"])

        if self.is_negative:
            self._ensure_sheet_exists(VALIDATION_SHEET["name"], VALIDATION_SHEET["index"])

    def _ensure_sheet_exists(self, name: str, index: int):
        if name not in self.wb.sheetnames:
            self.wb.create_sheet(name)

        sheet = self.wb[name]
        current_index = self.wb.index(sheet)

        if current_index != index:
            offset = index - current_index
            self.wb.move_sheet(sheet, offset=offset)

    def _populate_sheets(self):
        self._populate_library()
        self._populate_datasets()
        if self.is_negative:
            self._populate_validation()

    def _populate_library(self):
        ws = self.wb[LIBRARY_SHEET["name"]]
        for col_idx, header in enumerate(LIBRARY_SHEET["headers"], 1):
            ws.cell(row=1, column=col_idx, value=header)

        for cell_coord, value in LIBRARY_SHEET["defaults"].items():
            if not ws[cell_coord].value:
                ws[cell_coord] = value

    def _populate_datasets(self):
        ws = self.wb[DATASETS_SHEET["name"]]

        ws.delete_rows(1, ws.max_row)

        ws.append(DATASETS_SHEET["headers"])

        xpt_sheets = [s for s in self.wb.sheetnames if s.lower().endswith(".xpt")]

        for sheet_name in xpt_sheets:
            label = self.label_manager.get_label(sheet_name)
            ws.append([sheet_name, label])

    def _populate_validation(self):
        ws = self.wb[VALIDATION_SHEET["name"]]
        if ws.max_row == 1 and ws.cell(1, 1).value is None:
            for col_idx, header in enumerate(VALIDATION_SHEET["headers"], 1):
                ws.cell(row=1, column=col_idx, value=header)

    def _style_sheets(self):
        self._style_standard_sheet(LIBRARY_SHEET["name"])
        self._style_standard_sheet(DATASETS_SHEET["name"])

        if self.is_negative:
            self._style_standard_sheet(VALIDATION_SHEET["name"])

        self._style_xpt_sheets()

    def _style_standard_sheet(self, sheet_name: str):
        """Applies Bold Header style to a standard sheet."""
        if sheet_name not in self.wb.sheetnames:
            return

        ws = self.wb[sheet_name]
        for cell in ws[1]:
            cell.font = BOLD_FONT

    def _style_xpt_sheets(self):
        """Applies specific styling rules to .xpt sheets."""
        for sheet_name in self.wb.sheetnames:
            if not sheet_name.lower().endswith(".xpt"):
                continue

            ws = self.wb[sheet_name]

            for cell in ws[1]:
                cell.font = BOLD_FONT

            for row_idx in range(2, 5):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = LIGHT_FILL
                    cell.font = ITALIC_FONT


class Sanitiser:
    @staticmethod
    def sanitize_xlsx(file_path: str):
        path = Path(file_path)
        try:
            old_wb = openpyxl.load_workbook(path)

            new_wb = openpyxl.Workbook()
            new_wb.remove(new_wb.active)

            for sheet_name in old_wb.sheetnames:
                source = old_wb[sheet_name]
                target = new_wb.create_sheet(title=sheet_name)

                for row in source.iter_rows():
                    for cell in row:
                        new_cell = target.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.alignment = copy(cell.alignment)

            new_wb.save(path)

        except Exception as e:
            print(f"Error sanitizing {path.name}: {e}")


class InteractiveHandler:
    @staticmethod
    def select_cases(all_cases: List[dict]) -> List[dict]:
        print(f"\nFound {len(all_cases)} total test cases.")
        user_input = input("Enter specific case ID (e.g. CORE-000123/positive/01) or Press Enter for ALL: ").strip()

        if not user_input:
            return all_cases

        selected = [c for c in all_cases if f"{c['rule_id']}/{c['type']}/{c['case_id']}" == user_input]

        if not selected:
            print("No matching cases found.")
            return []
        return selected


def parse_arguments():
    parser = argparse.ArgumentParser(description="Excel Test Case Formatter")

    parser.add_argument("-r", "--rule", type=str, help="Rule ID (e.g., CORE-000123)")
    parser.add_argument("-tc", "--testcase", type=str, help="Test case sub-path (e.g., negative/01). Requires -r.")
    parser.add_argument("-all", "--all", action="store_true", help="Run on all test cases in all rules.")

    args = parser.parse_args()

    if args.testcase and not args.rule:
        parser.error("The -tc argument requires -r to be specified.")

    return args


def filter_cases_by_args(all_cases: List[dict], args) -> List[dict]:
    """Filters the full list of cases based on CLI arguments."""

    if args.all:
        return all_cases

    if args.rule:
        filtered = [c for c in all_cases if c["rule_id"] == args.rule]

        if args.testcase:
            filtered = [c for c in filtered if f"{c['type']}/{c['case_id']}" == args.testcase]

        if not filtered:
            msg = f"No cases found for Rule '{args.rule}'"
            if args.testcase:
                msg += f" and Case '{args.testcase}'"
            print(msg)
            sys.exit(1)

        return filtered

    return None


def main():
    args = parse_arguments()

    all_cases = FileManager.get_all_test_cases()
    if not all_cases:
        print("No test cases found in 'rules' directory.")
        sys.exit(1)

    label_mgr = LabelManager()
    label_mgr.scan_all_files(all_cases)

    selected_cases = filter_cases_by_args(all_cases, args)

    if selected_cases is None:
        selected_cases = InteractiveHandler.select_cases(all_cases)

    if not selected_cases:
        sys.exit(0)

    total_cases = len(selected_cases)
    print(f"Processing {total_cases} files...")
    success_count = 0

    for i, case in enumerate(selected_cases, 1):
        case_id_str = f"{case['rule_id']}/{case['type']}/{case['case_id']}"
        print(f"\r\033[K[{i}/{total_cases}] Processing: {case_id_str}", end="", flush=True)

        formatter = Formatter(case["path"], label_mgr)
        if formatter.format():
            success_count += 1

        Sanitiser.sanitize_xlsx(case["path"])

    print(f"\n\nComplete. {success_count}/{total_cases} formatted.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(0)
